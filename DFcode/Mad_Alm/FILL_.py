import pandas as pd
from openpyxl import load_workbook
import re
from tqdm import tqdm


def process_column_mapping(mapping_file_path, csv_file_path, excel_file_path, output_file_path, sheet_name='一元问题表'):
    """
    简单的列对列映射：将CSV的整列数据复制到Excel对应列

    Args:
        mapping_file_path: 映射文件路径
        csv_file_path: CSV文件路径（数据源）
        excel_file_path: Excel文件路径（目标表格）
        output_file_path: 输出文件路径
        sheet_name: Excel工作表名称
    """
    try:
        print("开始处理列映射...")

        # 步骤1: 读取映射文件
        print("正在读取映射规则...")
        mapping_rules = {}
        with open(mapping_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                cleaned = line.strip()
                if '-->' in cleaned and not cleaned.startswith('#'):
                    excel_col, csv_col = cleaned.split('-->', 1)
                    excel_col = excel_col.strip()
                    csv_col = csv_col.strip()
                    mapping_rules[excel_col] = csv_col

        print(f"读取到 {len(mapping_rules)} 条映射规则:")
        for excel_col, csv_col in mapping_rules.items():
            print(f"  Excel[{excel_col}] <-- CSV[{csv_col}]")

        # 步骤2: 读取CSV文件
        print("\n正在读取CSV文件...")
        csv_df = pd.read_csv(csv_file_path, encoding='utf-8', dtype=str)
        csv_df = csv_df.fillna('')  # 填充空值
        print(f"CSV文件包含 {len(csv_df)} 行数据")

        # 步骤3: 打开Excel文件
        print("\n正在读取Excel文件...")
        wb = load_workbook(excel_file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel文件中不存在工作表: {sheet_name}")

        ws = wb[sheet_name]

        # 步骤4: 创建Excel表头映射（支持模糊匹配）
        excel_headers = {}
        excel_headers_clean = {}  # 清理后的表头映射
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value:
                # 原始表头
                header_name = str(cell_value).strip()
                excel_headers[header_name] = col_idx

                # 清理后的表头（去除换行符、空格等）
                clean_header = re.sub(r'[\n\r\s]+', '', header_name)
                excel_headers_clean[clean_header] = col_idx

        print(f"Excel表头 ({len(excel_headers)} 列):")
        for header, col_idx in excel_headers.items():
            clean_header = re.sub(r'[\n\r\s]+', '', header)
            print(f"  列 {col_idx}: '{header}' -> 清理后: '{clean_header}'")

        # 步骤5: 先处理A列序号
        print("\n正在添加A列序号...")
        # 确保有足够的行数（基于CSV数据行数）
        data_rows = len(csv_df)

        # 在A列添加序号（从第2行开始，因为第1行是表头）
        for row_idx in range(2, 2 + data_rows):
            sequence_num = row_idx - 1  # 序号从1开始
            ws.cell(row=row_idx, column=1, value=sequence_num)

        print(f"  ✅ 在A列添加了 {data_rows} 个序号 (1-{data_rows})")

        # 步骤6: 执行列映射
        print("\n开始执行列映射...")
        successful_mappings = 0
        failed_mappings = []

        for excel_col, csv_col in mapping_rules.items():
            print(f"\n处理映射: Excel[{excel_col}] <-- CSV[{csv_col}]")

            # 检查Excel中是否存在目标列
            if excel_col not in excel_headers_clean:
                print(f"Excel中的列: { excel_headers_clean}")
                print(f"  ❌ Excel中未找到列: {excel_col}")
                failed_mappings.append(f"Excel中未找到列: {excel_col}")
                continue

            # 检查CSV中是否存在源列
            if csv_col not in csv_df.columns:
                print(f"  ❌ CSV中未找到列: {csv_col}")
                failed_mappings.append(f"CSV中未找到列: {csv_col}")
                continue

            # 获取Excel列索引
            excel_col_idx = excel_headers_clean[excel_col]

            # 复制数据
            csv_data = csv_df[csv_col].tolist()
            print(
                f"  📋 从CSV[{csv_col}]复制 {len(csv_data)} 行数据到Excel列{excel_col_idx}[{excel_col}]")

            # 写入Excel（从第2行开始，第1行是表头）
            for row_idx, value in enumerate(csv_data, start=2):
                if pd.isna(value):
                    value = ""
                else:
                    value = str(value).strip()
                ws.cell(row=row_idx, column=excel_col_idx, value=value)

            successful_mappings += 1
            print(f"  ✅ 成功映射: Excel[{excel_col}] <-- CSV[{csv_col}]")

        # 步骤6: 保存文件
        print(f"\n正在保存文件到: {output_file_path}")
        wb.save(output_file_path)

        # 输出统计信息
        print(f"\n" + "="*50)
        print(f"列映射处理完成!")
        print(f"映射规则总数: {len(mapping_rules)}")
        print(f"成功映射列数: {successful_mappings}")
        print(f"失败映射列数: {len(failed_mappings)}")

        if failed_mappings:
            print(f"\n失败的映射:")
            for failure in failed_mappings:
                print(f"  ❌ {failure}")

        print(f"\n结果已保存到: {output_file_path}")
        return True

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    # 设置文件路径
    mapping_file = r'DFcode\对应映射.txt'  # 修改为相对路径
    csv_file = r'DFcode\alm_issues合并结果.csv'  # 修改为相对路径
    excel_file = r'DFcode\测试.xlsx'  # 修改为相对路径
    output_file = r'DFcode\测试_更新后.xlsx'  # 修改为相对路径

    print("简单列映射处理工具")
    print("=" * 50)
    print("功能: 将CSV数据映射到Excel + 添加序号 + 处理Excel公式")
    print(f"映射文件: {mapping_file}")
    print(f"数据源(CSV): {csv_file}")
    print(f"目标文件(Excel): {excel_file}")
    print(f"输出文件: {output_file}")
    print("=" * 50)

    # 执行处理
    success = process_column_mapping(
        mapping_file_path=mapping_file,
        csv_file_path=csv_file,
        excel_file_path=excel_file,
        output_file_path=output_file,
        sheet_name='一元问题表'
    )

    if success:
        print("\n✅ 列映射处理成功完成！")
    else:
        print("\n❌ 列映射处理失败！")


if __name__ == "__main__":
    main()
